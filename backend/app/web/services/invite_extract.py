"""Candidate emails and roles out of an uploaded file — ports `server/services/inviteExtract.ts`.

Two paths. A spreadsheet is read column-wise, detecting which column holds the email
and which the role. An unstructured document (PDF, DOCX, text) is scanned by pattern,
and roles fall back to the batch role because there is no reliable way to pair them.

Deliberately permissive: **the recruiter's review table is the real gate.** This
returns rows with a `valid` flag rather than rejecting anything, so a malformed address
appears in the review UI where a human can fix it, instead of vanishing silently before
they ever see it. Nothing here creates an invite or sends anything.
"""

from __future__ import annotations

import csv
import io
import logging
import re

logger = logging.getLogger("web.invite_extract")

# Scanning prose for addresses. Permissive on purpose — see the module docstring.
EMAIL_PATTERN = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

# The shape check shown to the recruiter as valid/invalid. Not RFC-complete, and it
# does not need to be: delivery is the real test, and this only decides which rows the
# review table flags.
_VALID_EMAIL = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]{2,}$")

_EMAIL_HEADER = re.compile(r"\b(e-?mail|mail)\b", re.IGNORECASE)
_ROLE_HEADER = re.compile(
    r"\b(role|position|title|designation|profile|job)\b", re.IGNORECASE
)

CSV_SUFFIXES = (".csv", ".tsv")
XLSX_SUFFIXES = (".xlsx", ".xlsm")
# The legacy binary format. openpyxl cannot read it, and adding a dependency for a
# format Excel has not written by default since 2007 is not worth it — the recruiter
# is told to re-save instead.
LEGACY_EXCEL_SUFFIXES = (".xls",)


def is_valid_email(value: str) -> bool:
    return bool(_VALID_EMAIL.match((value or "").strip()))


def _is_spreadsheet(filename: str, content_type: str) -> bool:
    lowered = (filename or "").lower()
    mime = (content_type or "").lower()
    return (
        lowered.endswith(CSV_SUFFIXES + XLSX_SUFFIXES + LEGACY_EXCEL_SUFFIXES)
        or "spreadsheet" in mime
        or "excel" in mime
        or "csv" in mime
        or mime == "text/tab-separated-values"
    )


def _cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def rows_from_csv(data: bytes, *, delimiter: str | None = None) -> list[list[str]]:
    """A CSV or TSV as rows of strings.

    The delimiter is sniffed when not given, because recruiters export from tools that
    disagree — a semicolon-separated file from a European Excel would otherwise parse
    as one column per row and find nothing.
    """
    text = data.decode("utf-8-sig", "replace")
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
    return [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def rows_from_xlsx(data: bytes) -> list[list[str]]:
    """The first worksheet as rows of strings.

    `data_only=True` reads a formula's cached VALUE rather than its text: a sheet
    where the email column is `=CONCAT(...)` would otherwise yield the formula.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            return []
        sheet = workbook.worksheets[0]
        return [[_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def candidates_from_rows(
    rows: list[list[str]], fallback_role: str
) -> tuple[list[dict], bool]:
    """Rows → `{email, role}` pairs, plus whether a header row was detected.

    A row counts as a header when a cell NAMES an email column without being one
    itself — "Email" is a header, "ada@x.test" is data. Without that distinction a
    headerless file would lose its first candidate.
    """
    non_empty = [row for row in rows if any(_cell(cell) for cell in row)]
    if not non_empty:
        return [], False

    first = [_cell(cell) for cell in non_empty[0]]
    header_index = next(
        (
            index
            for index, cell in enumerate(first)
            if _EMAIL_HEADER.search(cell) and "@" not in cell
        ),
        -1,
    )

    if header_index >= 0:
        role_index = next(
            (index for index, cell in enumerate(first) if _ROLE_HEADER.search(cell)), -1
        )
        out = []
        for row in non_empty[1:]:
            cells = [_cell(cell) for cell in row]
            email = _at(cells, header_index)
            if not email:
                # The named column was blank for this row; fall back to any address in
                # it rather than dropping a candidate over a ragged export.
                email = next((cell for cell in cells if "@" in cell), "")
            role = (_at(cells, role_index) if role_index >= 0 else "") or fallback_role
            out.append({"email": email, "role": role})
        return out, True

    # No header: take the first address-looking cell per row. Roles cannot be mapped.
    out = [
        {
            "email": next((cell for cell in (_cell(c) for c in row) if "@" in cell), ""),
            "role": fallback_role,
        }
        for row in non_empty
    ]
    return out, False


def _at(cells: list[str], index: int) -> str:
    return cells[index] if 0 <= index < len(cells) else ""


def deduplicate(raw: list[dict], fallback_role: str) -> tuple[list[dict], int]:
    """Normalise and de-duplicate by lowercased email, preserving order.

    Case-insensitive because `Ada@x.test` and `ada@x.test` are one mailbox, and
    inviting both would send the same person two interviews with different ids.
    """
    seen: set[str] = set()
    duplicates = 0
    rows: list[dict] = []

    for entry in raw:
        email = (entry.get("email") or "").strip()
        if not email:
            continue
        key = email.lower()
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        rows.append(
            {
                "email": email,
                "role": (entry.get("role") or fallback_role).strip(),
                "valid": is_valid_email(email),
            }
        )

    return rows, duplicates


async def extract_candidates(
    data: bytes, *, content_type: str, filename: str, fallback_role: str
) -> dict:
    """`{rows, warnings}` for the recruiter's review step.

    Warnings are the honest part of this: every path that guessed something says so,
    because the recruiter is about to email real people.
    """
    warnings: list[str] = []
    raw: list[dict] = []

    if (filename or "").lower().endswith(LEGACY_EXCEL_SUFFIXES):
        return {
            "rows": [],
            "warnings": [
                "The old .xls format is not supported. Open the file and save it as "
                ".xlsx or .csv, then upload it again."
            ],
        }

    if _is_spreadsheet(filename, content_type):
        lowered = (filename or "").lower()
        try:
            if lowered.endswith(XLSX_SUFFIXES) or "excel" in (content_type or "").lower():
                rows = rows_from_xlsx(data)
            else:
                rows = rows_from_csv(data)
        except Exception as exc:  # noqa: BLE001 - a corrupt upload is the recruiter's problem to see
            logger.info("could not read spreadsheet %r: %s", filename, type(exc).__name__)
            return {
                "rows": [],
                "warnings": ["That spreadsheet could not be read. Try re-saving it as .csv."],
            }

        # No rows is NOT necessarily "no sheets": an empty CSV is simply empty, and
        # telling a recruiter their .csv "had no sheets" describes a file they did not
        # upload. Only the workbook path can genuinely lack a worksheet, and
        # `rows_from_xlsx` returns [] for that too — so both fall through to the
        # ordinary "no addresses found" message below, which is true either way.
        raw, headered = candidates_from_rows(rows, fallback_role)
        if rows and not headered:
            warnings.append(
                "No email/role header row detected — mapped the first email in each "
                "row and defaulted roles to the batch role."
            )
    else:
        from app.web.services import resume_text

        text = await resume_text.extract(
            data, content_type=content_type, filename=filename
        )
        found = EMAIL_PATTERN.findall(text)
        raw = [{"email": email, "role": fallback_role} for email in found]
        if found:
            warnings.append(
                "Unstructured file — emails were extracted by pattern and roles "
                "defaulted to the batch role. Please review carefully."
            )

    rows, duplicates = deduplicate(raw, fallback_role)
    if duplicates:
        plural = "" if duplicates == 1 else "s"
        warnings.append(f"{duplicates} duplicate email{plural} removed.")
    if not rows:
        warnings.append("No email addresses found in this file.")

    return {"rows": rows, "warnings": warnings}
